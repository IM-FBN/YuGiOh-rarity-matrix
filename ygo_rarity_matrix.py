import requests
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
import sqlite3
from pathlib import Path


COMMON_LIMIT = 3        # Sets a maxiumum of common cards 
FULL_REPORT=1           # if disabled only the raritys C,R,SR,UR,SCR, ULT are shown 
DB_DIR = str(Path.cwd())+os.sep+"expansions" # place EDOPro expensions folder to current directory or specify path  


#Some Comments are still in german language, i might change that some day



def find_main(alt_id_list):
    alt_id_list = [int(item) for item in alt_id_list]

    main_id = []
    SELECT_STATEMENT = """
        SELECT datas.id, datas.alias, texts.name 
        FROM datas
        INNER JOIN texts ON datas.id = texts.id 
        WHERE instr(texts.name, '(GOAT)') OR instr(texts.name, '(Pre-Errata)')
        ORDER BY texts.name
    """

    def find():
        databases = sorted(os.listdir(DB_DIR))
        databases = filter(lambda f: f.endswith(".cdb"), databases)


        hits = []
        encounteredIds = set()

        for db in databases:
            dbPath = (DB_DIR + os.sep +db)

            with sqlite3.connect(dbPath) as con:
                cursor = con.cursor()
                cursor.execute(SELECT_STATEMENT)
                rows = cursor.fetchall()
                for row in rows:
                    altId = row[0]
                    mainId = row[1]
                    name = row[2].strip()

                    if altId in encounteredIds:
                        continue

                    _type = "Unknown"
                    if "(GOAT)" in name:
                        _type = "GOAT"
                    elif "(Pre-Errata)" in name:
                        _type = "Pre-Errata"

                    hits.append({
                        "alt_id": altId,
                        "main_id": mainId,
                        "type": _type,
                        "name": name,
                        "db": db
                    })

                    encounteredIds.add(altId)

        def f(hit): return hit["name"]
        hits.sort(key=f)

        return hits


    if __name__ == "__main__":
        hits = find()
        cards = json.dumps(hits, indent=4, ensure_ascii=False)



    # Konvertiere den JSON-String in eine Liste von Dictionaries
    cards = json.loads(cards)

    # Liste mit den alt_id-Nummern, die überprüft werden sollen

    # Durchsuche die Kartenliste und gibt main_id aus, falls alt_id in der Liste vorkommt
    for card in cards:
        if card['alt_id'] in alt_id_list:
            main_id.append(str(card['main_id']))
    return main_id

def decklist_request(decklist):

    response_data = []
    error_list = []
    decklist = set(decklist)
    for id in decklist:
        response = requests.get('https://db.ygoprodeck.com/api/v7/cardinfo.php?'+'id='+str(id))
        response_json = response.json()
        # Skip parsing if there was an error with the request
        if "error" in response_json:
            error_list.append(id)
            continue
        response_data.append(response_json)

    
    error_list = set(error_list)
    main_id_list = find_main(error_list)

    # List all files in the current directory
    files = os.listdir(DB_DIR)

    # Check if any file has the extension ".cdb"
    if any(file.endswith(".cdb") for file in files):
        print(".cdb Files found")
        for id in main_id_list:
            response = requests.get('https://db.ygoprodeck.com/api/v7/cardinfo.php?'+'id='+str(id))
            response_json = response.json()
            # Skip parsing if there was an error with the request
            if "error" in response_json:
                print("Error with card ID {}: {}".format(id, response_json["error"]))
                error_list.append(id)
                continue
            response_data.append(response_json)

    else:
        print("There are no .cdb files in the specified directory.")
        #ERRORS IN NEUES TEXTFILE SCHREIBEN
        print(error_list)
        with open("error_cards.ydk", "w") as file:
            # Iteriere durch jede Karte in der error_list
            for card in error_list:
                # Schreibe die Karte in die Datei
                file.write(card + "\n")

    return response_data


def parse_card_data(card_data):
    # Erstelle eine leere Datenstruktur, um die Ergebnisse zu speichern
    result = {}

    # Gehe durch jedes Element in der "data" Liste
    for info in card_data["data"]:
        # Erstelle eine leere Liste, um die Seltenheitswerte zu speichern
        rarities = {}

        # Gehe durch jedes Element in der "card_sets" Liste
        for card_set in info["card_sets"]:
            # Überprüfe, ob die Seltenheit bereits in der "rarities" Liste enthalten ist
            if card_set["set_rarity"] not in rarities:
                # Wenn nicht, füge sie der Liste hinzu
                rarities[card_set["set_rarity"]] = [card_set["set_code"].split("-")[0]+" "+card_set["set_price"]+"$"]
            else:
                rarities[card_set["set_rarity"]].append(card_set["set_code"].split("-")[0]+" "+card_set["set_price"]+"$")

        # Füge den Kartenname und die Liste der Seltenheitswerte dem Ergebnis hinzu
        result.update({"name": info["name"], "rarity": rarities})

    return result





def data_to_excel(parsed_data, FULL_REPORT=1, ALTERNATE_BG_COLOR=True):
    # Kartennamen alphabetisch sortieren
    data_sorted = sorted(parsed_data, key=lambda card: card["name"])

    # Spaltennamen für die Tabelle erstellen
    rarities = ["Common", "Rare", "Super Rare", "Ultra Rare", "Secret Rare", "Ultimate Rare"]
    if FULL_REPORT == 1:
        for card in data_sorted:
            for rarity in card["rarity"].keys():
                if rarity not in rarities:
                    rarities.append(rarity)
    column_names = ["Name"] + rarities

    # Excel-Arbeitsmappe erstellen
    workbook = Workbook()
    sheet = workbook.active

    # Spaltennamen in die Tabelle schreiben
    for i, name in enumerate(column_names):
        cell = sheet.cell(row=1, column=i+1, value=name)
        # Färbe jede zweite Spalte mit einer hellgrauen Hintergrundfarbe, wenn ALTERNATE_BG_COLOR=True
        if ALTERNATE_BG_COLOR and i % 2 == 1:
            fill = PatternFill(start_color='E3E3E3', end_color='E3E3E3', fill_type='solid')
            cell.fill = fill

    # Daten in die Tabelle schreiben
    #Fixiere die erste Spalte
    sheet.freeze_panes = 'B2'

    for i, card in enumerate(data_sorted):
        row = [card["name"]]
        for j, rarity in enumerate(rarities):
            sets = card["rarity"].get(rarity, [])
            if rarity == "Common":
                # Begrenze die Anzahl der Einträge für "Common"
                if len(sets) > COMMON_LIMIT:
                    row.append("\n".join(sets[:COMMON_LIMIT]) + "\n ({} weitere)".format(len(sets) - COMMON_LIMIT))
                else:
                    row.append("\n".join(sets))
            else:
                row.append("\n".join(sets))
        for j, value in enumerate(row):
            cell = sheet.cell(row=i+2, column=j+1, value=value)
            # Set alignment of cell to top and left
            cell.alignment = Alignment(horizontal="left", vertical="top")
            # Auto-size column width based on cell content
            column_letter = get_column_letter(j+1)
            column_dimension = sheet.column_dimensions[column_letter]
            column_dimension.auto_size = True
            max_length = 0
            # Färbe jede zweite Spalte mit einer hellgrauen Hintergrundfarbe, wenn ALTERNATE_BG_COLOR=True
            # Set top border of cell
            border = Border(top=Side(border_style="thin", color="000000"))
            cell.border = border


            if ALTERNATE_BG_COLOR and j % 2 == 1:
                fill = PatternFill(start_color='E3E3E3', end_color='E3E3E3', fill_type='solid')
                cell.fill = fill

    # Enable text wrap for all cells
    for row in sheet.rows:
        for cell in row:
            cell.alignment = Alignment(wrapText=True, horizontal="right", vertical="top")

    # Excel-Datei speichern
    workbook.save("rarity_matrix.xlsx")
    return "saved"

ALTERNATE_BG_COLOR=1
decklist = []

for file_name in os.listdir():
    if file_name.endswith(".ydk"):
        with open(file_name) as file:
            for line in file:
                entry = line.strip()
                if entry[0].isdigit():
                    decklist.append(entry)


decklist_data = decklist_request(decklist)    
parsed_data =[]
for card_data in decklist_data:
    parsed_data.append(parse_card_data(card_data))
data_to_excel(parsed_data, FULL_REPORT, ALTERNATE_BG_COLOR)
print("FILE CREATED - FINISHED")
    
